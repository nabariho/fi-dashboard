#!/usr/bin/env python3
"""
Import Bankinter 2025 expense data into FI Dashboard cashflow format.

Data sources:
  1. Account movements XLSX — salary income, direct debits, Bizum, donations
  2. Credit card PDFs (12 months) — individual card transactions

Output: cashflow_import_2025.json with cashflowEntries, cashflowCategories,
        and cashflowSubcategories ready for dashboard import.

Rules:
  - Card transactions are categorized by merchant pattern matching
  - Account XLSX "RECIBO VISA CLASICA" entries are EXCLUDED (already in card PDFs)
  - Broker transfers + Revolut/BBVA/joint account = 'transfer' classification (not spending)
  - Refunds (ANUL.) reduce the corresponding category total
  - Amounts are aggregated per month + category + subcategory
  - Categories carry a 'classification' field: 'spending' (real expenses) or 'transfer'
"""

import json
import re
import os
from collections import defaultdict
from datetime import datetime

import openpyxl
import pdfplumber

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INBOX_DIR = os.path.join(BASE_DIR, 'inbox', 'bankinter')
XLSX_PATH = os.path.join(INBOX_DIR, 'account', 'Movimientos_2025.xlsx')
CARD_DIR = os.path.join(INBOX_DIR, 'card')

# ─── Category & Subcategory Taxonomy ────────────────────────────────────

EXPENSE_CATEGORIES = {
    'Food Delivery': {'id': 'expense_food-delivery', 'sort': 1},
    'Restaurants': {'id': 'expense_restaurants', 'sort': 2},
    'Groceries': {'id': 'expense_groceries', 'sort': 3},
    'Transport': {'id': 'expense_transport', 'sort': 4},
    'Fuel': {'id': 'expense_fuel', 'sort': 5},
    'Shopping': {'id': 'expense_shopping', 'sort': 6},
    'Subscriptions': {'id': 'expense_subscriptions', 'sort': 7},
    'Health': {'id': 'expense_health', 'sort': 8},
    'Car': {'id': 'expense_car', 'sort': 9},
    'Travel': {'id': 'expense_travel', 'sort': 10},
    'Education': {'id': 'expense_education', 'sort': 11},
    'Insurance': {'id': 'expense_insurance', 'sort': 12},
    'Donations': {'id': 'expense_donations', 'sort': 13},
    'Entertainment': {'id': 'expense_entertainment', 'sort': 14},
    'Gifts': {'id': 'expense_gifts', 'sort': 15},
    'Utilities': {'id': 'expense_utilities', 'sort': 16},
    'Bank Fees': {'id': 'expense_bank-fees', 'sort': 17},
    'Leisure': {'id': 'expense_leisure', 'sort': 18},
    'Housing': {'id': 'expense_housing', 'sort': 19},
    'Taxes': {'id': 'expense_taxes', 'sort': 20},
    'Investing': {'id': 'expense_investing', 'sort': 21, 'classification': 'transfer'},
    'Internal Transfer': {'id': 'expense_internal-transfer', 'sort': 22, 'classification': 'transfer'},
    'Other': {'id': 'expense_other', 'sort': 99},
}

INCOME_CATEGORIES = {
    'Salary': {'id': 'income_salary', 'sort': 1},
    'Bonus': {'id': 'income_bonus', 'sort': 2},
    'Bizum': {'id': 'income_bizum', 'sort': 3},
    'Other': {'id': 'income_other', 'sort': 4},
}

SUBCATEGORIES = {
    # Food Delivery
    'expense_food-delivery': {
        'Uber Eats': {'id': 'expense_food-delivery_uber-eats', 'sort': 1},
        'Glovo': {'id': 'expense_food-delivery_glovo', 'sort': 2},
    },
    # Transport
    'expense_transport': {
        'Uber Rides': {'id': 'expense_transport_uber-rides', 'sort': 1},
        'Taxi': {'id': 'expense_transport_taxi', 'sort': 2},
        'Public Transport': {'id': 'expense_transport_public-transport', 'sort': 3},
        'Parking': {'id': 'expense_transport_parking', 'sort': 4},
    },
    # Shopping
    'expense_shopping': {
        'El Corte Ingles': {'id': 'expense_shopping_el-corte-ingles', 'sort': 1},
        'Amazon': {'id': 'expense_shopping_amazon', 'sort': 2},
        'Decathlon': {'id': 'expense_shopping_decathlon', 'sort': 3},
        'Other': {'id': 'expense_shopping_other', 'sort': 99},
    },
    # Subscriptions
    'expense_subscriptions': {
        'Apple': {'id': 'expense_subscriptions_apple', 'sort': 1},
        'Gamma': {'id': 'expense_subscriptions_gamma', 'sort': 2},
        'Amazon Prime': {'id': 'expense_subscriptions_amazon-prime', 'sort': 3},
        'Godaddy': {'id': 'expense_subscriptions_godaddy', 'sort': 4},
        'Other': {'id': 'expense_subscriptions_other', 'sort': 99},
    },
    # Health
    'expense_health': {
        'Dental': {'id': 'expense_health_dental', 'sort': 1},
        'Pharmacy': {'id': 'expense_health_pharmacy', 'sort': 2},
    },
    # Leisure
    'expense_leisure': {
        'Shared Events': {'id': 'expense_leisure_shared-events', 'sort': 1},
    },
    # Investing
    'expense_investing': {
        'Interactive Brokers': {'id': 'expense_investing_interactive-brokers', 'sort': 1},
        'Indexa Capital': {'id': 'expense_investing_indexa-capital', 'sort': 2},
        'Trade Republic': {'id': 'expense_investing_trade-republic', 'sort': 3},
    },
    # Internal Transfer
    'expense_internal-transfer': {
        'Revolut': {'id': 'expense_internal-transfer_revolut', 'sort': 1},
        'BBVA': {'id': 'expense_internal-transfer_bbva', 'sort': 2},
        'Joint Account': {'id': 'expense_internal-transfer_joint-account', 'sort': 3},
    },
    # Car
    'expense_car': {
        'Repair': {'id': 'expense_car_repair', 'sort': 1},
        'Rental': {'id': 'expense_car_rental', 'sort': 2},
    },
}

# ─── Merchant → Category Mapping (card transactions) ───────────────────

# Order matters: first match wins. Patterns are case-insensitive.
CARD_RULES = [
    # Revolut top-ups — internal transfer, not spending
    (r'REVOLUT\*\*7530', 'Internal Transfer', 'Revolut'),

    # Food Delivery
    (r'UBER \*EATS|UBER\*EATS', 'Food Delivery', 'Uber Eats'),
    (r'GLOVO|PAYPAL \*GLOVO', 'Food Delivery', 'Glovo'),

    # Transport — Uber rides (not Eats)
    (r'UBER[, ]|UBR\*', 'Transport', 'Uber Rides'),
    (r'TAXI|LM \d+', 'Transport', 'Taxi'),
    (r'Pontevedr|Vigo|Baiona', 'Transport', 'Public Transport'),
    (r'PARKIA', 'Transport', 'Parking'),

    # Groceries
    (r'HIPERDINO|HD ', 'Groceries', None),
    (r'MERCADONA', 'Groceries', None),
    (r'SPAR', 'Groceries', None),
    (r'RINCON DEL PAN', 'Groceries', None),

    # Fuel
    (r'ESTACION BP|E S SAN LORENZO|SHELL', 'Fuel', None),

    # Restaurants / Dining
    (r'CHULETON|RINCON DE PAQUI|CAFE REGINA|CAFETERIA|CORIASSO|100 MONTADITOS|MCDONALD|KFC|SPAZIO|CASA GOYO|MOCANERA|BAR BOXES|GUACHINCHE|LA PATINETA|RESTAURANTE|ASADOR|EL JARDIN|TASQUITA|CASA JUANCRI|LA PATA CALIENTE|EL CARDONAL|THE BLOCK|SHINTORI|PASTELERIA|TASCA LA LONJA|CAFE REGINA', 'Restaurants', None),

    # Shopping — El Corte Inglés
    (r'ELCORTEINGL|NICOLAS ESTEVAN', 'Shopping', 'El Corte Ingles'),
    # Shopping — Amazon
    (r'AMAZON|AMZN', 'Shopping', 'Amazon'),
    # Shopping — Decathlon
    (r'DECATHLON', 'Shopping', 'Decathlon'),
    # Shopping — Other stores
    (r'ALEHOP|MULTIPRECIOS|DANIHACE|PINTADERA|FOTODIGITAL|FUNDGRUBE', 'Shopping', 'Other'),

    # Subscriptions
    (r'APPLE\.COM/BILL', 'Subscriptions', 'Apple'),
    (r'GAMMA\.APP', 'Subscriptions', 'Gamma'),
    (r'GODADDY', 'Subscriptions', 'Godaddy'),
    (r'COURSERA', 'Education', None),
    (r'Prime Video|primevideo', 'Subscriptions', 'Amazon Prime'),
    (r'HUMBLE', 'Entertainment', None),

    # Health
    (r'GIL TAFACHE|CLIN\. DENTAL|ESTUDIO DENTA', 'Health', 'Dental'),
    (r'FCIA|FARMACIA', 'Health', 'Pharmacy'),

    # Car
    (r'AUTOS YANEZ', 'Car', 'Repair'),
    (r'CICAR|BENICOCHES', 'Car', 'Rental'),

    # Travel
    (r'HOTEL|HEYMONDO|BROADWAY', 'Travel', None),

    # Education
    (r'HOTMART|HARVBUSREV', 'Education', None),

    # Entertainment
    (r'TULOTERO', 'Entertainment', None),

    # Gifts
    (r'FLEURO', 'Gifts', None),

    # Other — divvy diary (app)
    (r'DIVVYDIARY', 'Subscriptions', 'Other'),

    # US Customs (duty)
    (r'USCUSTOMSBO', 'Shopping', 'Other'),

    # Guanxe (online grocery/shop)
    (r'GUANXE', 'Shopping', 'Other'),

    # Tanatorio (funeral services)
    (r'TANATORIO', 'Other', None),

    # Cajero (ATM)
    (r'CAJERO', 'Bank Fees', None),

    # Sagulpa (parking meters)
    (r'SAGULPA', 'Transport', 'Parking'),

    # Card maintenance fee
    (r'Comisi.n mantenimiento', 'Bank Fees', None),

    # International commission fees
    (r'COM\. USO REDES', 'Bank Fees', None),

    # Catch-all for PAYPAL merchants not matched above
    (r'PAYPAL \*MADMUSL', 'Shopping', 'Other'),

    # Unmatched merchants from data review
    (r'ANTONIO MARTIN MAROTO', 'Transport', 'Taxi'),
    (r'FRANCISCO JAVIER PEREZ', 'Other', None),
    (r'REST\. MIRADOR|MIRADOR SANTA LUCIA', 'Restaurants', None),
]

# ─── Account XLSX Rules ────────────────────────────────────────────────

# Patterns for XLSX "Descripción" field
XLSX_INCOME_RULES = [
    (r'NOMI.*AGILE MONKEY|THE AGILE MONKEY|NOMINA', 'Salary', None),
    (r'BIZUM', 'Bizum', None),
]

XLSX_EXPENSE_RULES = [
    # === TRANSFER categories (not real spending) ===
    # Investment transfers
    (r'interactive.?bro', 'Investing', 'Interactive Brokers'),
    (r'Indexa', 'Investing', 'Indexa Capital'),
    (r'Trade Republic', 'Investing', 'Trade Republic'),
    # Internal transfers between own accounts
    (r'Revolut', 'Internal Transfer', 'Revolut'),
    (r'Mi BBVA|Bbva telefono', 'Internal Transfer', 'BBVA'),
    (r'Eva y Ruben', 'Internal Transfer', 'Joint Account'),

    # === SPENDING categories (real expenses) ===
    # Housing — rent payments to landlord (various transfer formats)
    (r'Esteban Betanc', 'Housing', None),
    # Housing — property management
    (r'Sanper Gestiones', 'Housing', None),
    # Insurance direct debits
    (r'MAPFRE|SEGURO', 'Insurance', None),
    # Utilities (phone, internet, etc.)
    (r'VODAFONE|MOVISTAR|ORANGE|ENDESA|IBERDROLA|NATURGY|CANAL DE ISABEL', 'Utilities', None),
    # Donations
    (r'DONACION|CRUZ ROJA|UNICEF|MEDICOS SIN|SAVE THE|GREENPEACE|ACNUR|CARITAS|FUNDACION|AMNISTIA|ASOCIACION ESPANOLA CON', 'Donations', None),
    # Taxes
    (r'TRIBUTO|VALORA GESTION TRIBUTAR', 'Taxes', None),
    # Bizum expenses — shared events / leisure
    (r'BIZUM', 'Leisure', 'Shared Events'),
    # Bank fees
    (r'COMISION|COM\.', 'Bank Fees', None),
]

# Patterns to EXCLUDE from XLSX (not tracked at all)
XLSX_EXCLUDE_PATTERNS = [
    r'RECIBO VISA',                   # Card bill — already detailed in card PDFs
    r'TRASPASO',                      # Internal bank transfers (no external account)
    r'A CUENTA .* BANKINTER',         # Internal Bankinter transfers
    r'V\.M\.Dia-Ruben Dominguez',     # Self-transfer
]


def slugify(s):
    return re.sub(r'[^a-z0-9-]', '', re.sub(r'\s+', '-', s.strip().lower()))


# ─── Parse Card PDFs ───────────────────────────────────────────────────

def parse_card_pdf(pdf_path):
    """Extract transactions from a Bankinter card statement PDF."""
    transactions = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 4:
                        continue
                    # Row format: [date, card_digits, concept, cargo, abono]
                    date_str = (row[0] or '').strip()
                    # Must match dd/mm/yyyy
                    if not re.match(r'\d{2}/\d{2}/\d{4}$', date_str):
                        continue
                    concept = (row[2] or '').strip()
                    if not concept:
                        continue

                    # Parse amounts (Spanish format: 1.234,56)
                    cargo_str = (row[3] or '').strip().replace('.', '').replace(',', '.')
                    abono_str = (row[4] or '').strip().replace('.', '').replace(',', '.') if len(row) > 4 else ''

                    cargo = float(cargo_str) if cargo_str else 0
                    abono = float(abono_str) if abono_str else 0

                    date = datetime.strptime(date_str, '%d/%m/%Y')
                    month = date.strftime('%Y-%m')

                    transactions.append({
                        'date': date,
                        'month': month,
                        'concept': concept,
                        'cargo': cargo,
                        'abono': abono,
                        'source': 'card',
                    })
    return transactions


def categorize_card_tx(concept, is_refund=False):
    """Match a card transaction concept to a category/subcategory."""
    for pattern, category, subcategory in CARD_RULES:
        if re.search(pattern, concept, re.IGNORECASE):
            return category, subcategory
    # Fallback
    print(f"  UNMATCHED card tx: {concept}")
    return 'Other', None


# ─── Parse Account XLSX ────────────────────────────────────────────────

def parse_xlsx(xlsx_path):
    """Extract transactions from Bankinter account movements XLSX."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    transactions = []

    # Find header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        values = [str(cell.value or '').strip().lower() for cell in row]
        if 'fecha contable' in values or 'fecha' in values:
            header_row = row_idx
            headers = values
            break

    if not header_row:
        # Try row 1 as header
        header_row = 1
        headers = [str(cell.value or '').strip().lower() for cell in ws[1]]

    # Find column indices
    date_col = next((i for i, h in enumerate(headers) if 'fecha' in h and 'valor' not in h), 0)
    desc_col = next((i for i, h in enumerate(headers) if 'desc' in h or 'concepto' in h), 2)
    amount_col = next((i for i, h in enumerate(headers) if 'importe' in h), 3)

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        date_val = row[date_col]
        desc = str(row[desc_col] or '').strip()
        amount_val = row[amount_col]

        if not date_val or not desc:
            continue

        # Parse date
        if isinstance(date_val, datetime):
            date = date_val
        elif isinstance(date_val, str):
            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                try:
                    date = datetime.strptime(date_val.strip(), fmt)
                    break
                except ValueError:
                    continue
            else:
                continue
        else:
            continue

        # Parse amount
        if isinstance(amount_val, (int, float)):
            amount = float(amount_val)
        elif isinstance(amount_val, str):
            amount = float(amount_val.replace('.', '').replace(',', '.'))
        else:
            continue

        month = date.strftime('%Y-%m')
        transactions.append({
            'date': date,
            'month': month,
            'concept': desc,
            'amount': amount,  # positive = income, negative = expense
            'source': 'account',
        })

    return transactions


# ─── Main Processing ───────────────────────────────────────────────────

def main():
    # === 1. Parse all card PDFs ===
    print("=== Parsing card PDFs ===")
    all_card_txs = []
    for month_num in range(1, 13):
        pdf_name = f'bankinter_card_2025_{month_num:02d}.pdf'
        pdf_path = os.path.join(CARD_DIR, pdf_name)
        if os.path.exists(pdf_path):
            txs = parse_card_pdf(pdf_path)
            print(f"  {pdf_name}: {len(txs)} transactions")
            all_card_txs.extend(txs)
        else:
            print(f"  {pdf_name}: NOT FOUND")

    print(f"\nTotal card transactions: {len(all_card_txs)}")

    # === 2. Categorize card transactions ===
    print("\n=== Categorizing card transactions ===")
    # month -> category -> subcategory -> amount
    cashflow = defaultdict(lambda: defaultdict(float))

    for tx in all_card_txs:
        # Skip transactions from previous year (billing period overlap)
        if not tx['month'].startswith('2025'):
            continue
        is_refund = 'ANUL.' in tx['concept'] or tx['abono'] > 0
        category, subcategory = categorize_card_tx(tx['concept'], is_refund)

        if category is None:
            # Excluded (e.g., Revolut top-up)
            continue

        amount = tx['cargo'] if tx['cargo'] > 0 else -tx['abono']
        if is_refund and tx['abono'] > 0:
            amount = -tx['abono']

        key = (category, subcategory or '')
        cashflow[tx['month']][key] += amount

    # === 3. Parse account XLSX ===
    print("\n=== Parsing account XLSX ===")
    xlsx_txs = parse_xlsx(XLSX_PATH)
    print(f"Total XLSX transactions: {len(xlsx_txs)}")

    # === 4. Categorize XLSX transactions ===
    print("\n=== Categorizing XLSX transactions ===")
    income_data = defaultdict(lambda: defaultdict(float))  # month -> category -> amount
    skipped = 0
    unmatched_expenses = []

    for tx in xlsx_txs:
        desc = tx['concept']
        amount = tx['amount']

        # Check if excluded
        excluded = False
        for pattern in XLSX_EXCLUDE_PATTERNS:
            if re.search(pattern, desc, re.IGNORECASE):
                excluded = True
                break
        if excluded:
            skipped += 1
            continue

        # Income (positive amounts)
        if amount > 0:
            matched = False
            for pattern, category, subcategory in XLSX_INCOME_RULES:
                if re.search(pattern, desc, re.IGNORECASE):
                    income_data[tx['month']][(category, subcategory or '')] += amount
                    matched = True
                    break
            if not matched:
                # Other income
                income_data[tx['month']][('Other', '')] += amount
            continue

        # Expense (negative amounts)
        abs_amount = abs(amount)
        matched = False
        for pattern, category, subcategory in XLSX_EXPENSE_RULES:
            if re.search(pattern, desc, re.IGNORECASE):
                key = (category, subcategory or '')
                cashflow[tx['month']][key] += abs_amount
                matched = True
                break

        if not matched:
            unmatched_expenses.append(tx)

    print(f"  Skipped (excluded): {skipped}")
    print(f"  Unmatched expenses: {len(unmatched_expenses)}")
    for tx in unmatched_expenses:
        print(f"    {tx['month']} | {tx['amount']:>10.2f} | {tx['concept']}")

    # === 5. Build output ===
    print("\n=== Building cashflow entries ===")

    # Build categories list
    categories = []
    for name, info in EXPENSE_CATEGORIES.items():
        cat = {
            'category_id': info['id'],
            'type': 'expense',
            'name': name,
            'active': True,
            'sort_order': info['sort'],
            'classification': info.get('classification', 'spending'),
        }
        categories.append(cat)
    for name, info in INCOME_CATEGORIES.items():
        categories.append({
            'category_id': info['id'],
            'type': 'income',
            'name': name,
            'active': True,
            'sort_order': info['sort'],
        })

    # Build subcategories list
    subcategories = []
    for cat_id, subs in SUBCATEGORIES.items():
        for name, info in subs.items():
            subcategories.append({
                'subcategory_id': info['id'],
                'category_id': cat_id,
                'name': name,
                'active': True,
                'sort_order': info['sort'],
            })

    # Build entries
    entries = []
    used_ids = set()

    def make_entry_id(month, entry_type, cat_id, subcat_id=None):
        base = f"{month}_{entry_type}_{cat_id}"
        if subcat_id:
            base += f"_{subcat_id}"
        eid = base
        n = 2
        while eid in used_ids:
            eid = f"{base}__{n}"
            n += 1
        used_ids.add(eid)
        return eid

    # Expense entries from cashflow aggregation
    for month in sorted(cashflow.keys()):
        for (category, subcategory), amount in sorted(cashflow[month].items()):
            if abs(amount) < 0.01:
                continue
            cat_info = EXPENSE_CATEGORIES.get(category)
            if not cat_info:
                print(f"  WARNING: unknown category '{category}'")
                continue

            cat_id = cat_info['id']
            subcat_id = None
            subcat_name = ''
            if subcategory and cat_id in SUBCATEGORIES:
                sub_info = SUBCATEGORIES[cat_id].get(subcategory)
                if sub_info:
                    subcat_id = sub_info['id']
                    subcat_name = subcategory

            entry_id = make_entry_id(month, 'expense', cat_id, subcat_id)
            entry = {
                'entry_id': entry_id,
                'month': month,
                'type': 'expense',
                'category': category,
                'category_id': cat_id,
                'amount': round(amount, 2),
                'notes': 'Imported from Bankinter 2025',
            }
            if subcat_id:
                entry['subcategory_id'] = subcat_id
                entry['subcategory'] = subcat_name
            else:
                entry['subcategory_id'] = None
                entry['subcategory'] = ''
            entries.append(entry)

    # Income entries
    for month in sorted(income_data.keys()):
        for (category, subcategory), amount in sorted(income_data[month].items()):
            if abs(amount) < 0.01:
                continue
            cat_info = INCOME_CATEGORIES.get(category)
            if not cat_info:
                print(f"  WARNING: unknown income category '{category}'")
                continue

            cat_id = cat_info['id']
            entry_id = make_entry_id(month, 'income', cat_id)
            entries.append({
                'entry_id': entry_id,
                'month': month,
                'type': 'income',
                'category': category,
                'category_id': cat_id,
                'subcategory_id': None,
                'subcategory': '',
                'amount': round(amount, 2),
                'notes': 'Imported from Bankinter 2025',
            })

    # Sort entries by month desc, then entry_id
    entries.sort(key=lambda e: (-int(e['month'].replace('-', '')), e['entry_id']))

    # === 6. Summary ===
    print(f"\nTotal entries: {len(entries)}")
    print("\n--- Monthly Summary ---")
    # Build a set of transfer category IDs for splitting
    transfer_cat_ids = set()
    for name, info in EXPENSE_CATEGORIES.items():
        if info.get('classification') == 'transfer':
            transfer_cat_ids.add(info['id'])

    month_totals = defaultdict(lambda: {'income': 0, 'spending': 0, 'transfers': 0})
    for e in entries:
        if e['type'] == 'income':
            month_totals[e['month']]['income'] += e['amount']
        elif e.get('category_id') in transfer_cat_ids:
            month_totals[e['month']]['transfers'] += e['amount']
        else:
            month_totals[e['month']]['spending'] += e['amount']

    for month in sorted(month_totals.keys()):
        t = month_totals[month]
        savings = t['income'] - t['spending']
        rate = (savings / t['income'] * 100) if t['income'] > 0 else 0
        print(f"  {month}  Income: {t['income']:>8.2f}  Spending: {t['spending']:>8.2f}  "
              f"Transfers: {t['transfers']:>8.2f}  Net Savings: {savings:>8.2f}  Rate: {rate:>5.1f}%")

    total_income = sum(t['income'] for t in month_totals.values())
    total_spending = sum(t['spending'] for t in month_totals.values())
    total_transfers = sum(t['transfers'] for t in month_totals.values())
    print(f"\n  TOTAL   Income: {total_income:>8.2f}  Spending: {total_spending:>8.2f}  "
          f"Transfers: {total_transfers:>8.2f}  Net Savings: {total_income - total_spending:>8.2f}")

    print("\n--- Category Breakdown ---")
    cat_totals = defaultdict(float)
    for e in entries:
        if e['type'] == 'expense':
            cat_totals[e['category']] += e['amount']
    for cat, total in sorted(cat_totals.items(), key=lambda x: -x[1]):
        print(f"  {cat:<20s}  {total:>8.2f}")

    # === 7. Write output ===
    output = {
        'cashflowCategories': categories,
        'cashflowSubcategories': subcategories,
        'cashflowEntries': entries,
    }

    output_path = os.path.join(BASE_DIR, 'cashflow_import_2025.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False, default=str)

    print(f"\nOutput written to: {output_path}")
    print("Import this file into the dashboard via Admin > Cash Flow > Import")


if __name__ == '__main__':
    main()
